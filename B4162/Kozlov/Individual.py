#Построение графика с диаграммами зависимости предела прочности плёнок от их сырья и концентрации. 
#Это относится к теме моей дипломной работы.

import pandas as pd 
import matplotlib as mpl 
import matplotlib.pyplot as plt 
import openpyxl as xl

file = xl.load_workbook('Данные.xlsx')
def writing(file):
    table = file['Лист1']
    string = input ('Введите новый материал: ')
    data = string.split(' ')
    data[1] = float(data[1])
    data[2] = float(data[2])
    i = 1
    while True:
        if not table.cell(row=i,column=2).value:
            table.cell(row=i,column=1).value = data[0]
            table.cell(row=i,column=2).value = data[1]
            table.cell(row=i,column=3).value = data[2]
            break
        i += 1
    file.save('Данные.xlsx')

table = pd.read_excel('Данные.xlsx') 
data_names = table.T._info_axis.values[:]
data_values = table.values[:, 1] 
y1 = table.values[:,1] 

dpi = 80 
fig = plt.figure(dpi = dpi, figsize = (512 / dpi, 384 / dpi) ) 
mpl.rcParams.update({'font.size': 10}) 

plt.title('Предел прочности при прокалывании σ (МПа)') 

ax = plt.axes() 
ax.yaxis.grid(True, zorder = 1) 

xs = range(len(data_names)) 

plt.bar([x + 0.05 for x in xs], [ d * 0.9 for d in data_values], 
width = 0.2, color = 'red', alpha = 0.9, label = '0.5 %', 
zorder = 2) 
plt.bar([x + 0.3 for x in xs], data_values, 
width = 0.2, color = 'blue', alpha = 0.9, label = '0.75 %', 
zorder = 2) 
plt.xticks(xs, data_names) 

fig.autofmt_xdate(rotation = 25) 

plt.legend(loc='upper right') 
fig.savefig('bars.png')
